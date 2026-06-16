"""Export des résultats en CSV (toujours) et Excel (si openpyxl est installé)."""
from __future__ import annotations

import csv
from typing import Any, Dict, List, Sequence

# Colonnes ordonnées pour une lecture humaine (le prix croissant en tête de logique).
PRODUCT_COLUMNS: Sequence[str] = (
    "price_min",
    "price_max",
    "currency",
    "price_min_xaf",
    "gauge_type",
    "title",
    "supplier_name",
    "min_order_qty",
    "unit",
    "price_raw",
    "product_url",
    "supplier_url",
    "supplier_id",
    "product_id",
    "source",
    "fetched_at",
)

SUPPLIER_COLUMNS: Sequence[str] = (
    "name",
    "location",
    "rating",
    "years_on_platform",
    "url",
    "supplier_id",
    "source",
)


def _ordered(row: Dict[str, Any], columns: Sequence[str]) -> List[Any]:
    return [row.get(col, "") for col in columns]


def export_csv(rows: List[Dict[str, Any]], columns: Sequence[str], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(columns)
        for row in rows:
            writer.writerow(_ordered(row, columns))


def export_xlsx(rows: List[Dict[str, Any]], columns: Sequence[str], path: str, sheet_name: str = "Sheet1") -> bool:
    """Exporte en .xlsx. Renvoie False (sans lever d'erreur) si openpyxl manque."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(_ordered(row, columns))

    # Largeur de colonne approximative basée sur la longueur de l'en-tête.
    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(48, len(col) + 6))

    wb.save(path)
    return True
